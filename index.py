import openpyxl
import pandas as pd
wb=openpyxl.load_workbook("Python Assignment.xlsx")
sh1=wb['(Input) User IDs']
sh2=wb['(Input) Rigorbuilder RAW']
ls=list(set(sh1.cell(row=i, column=6).value for i in range(12, 33)))
dict={}
for user_id in ls:
    dict[user_id] = []
for i in range(12, 33):
    user_id = sh1.cell(row=i, column=6).value
    statement = sh1.cell(row=i, column=7).value
    dict[user_id].append(statement)
dict2={}
t0=dict[ls[0]];t1=dict[ls[1]];t2=dict[ls[2]];t3=dict[ls[3]];t4=dict[ls[4]];t5=dict[ls[5]];t6=dict[ls[6]];t7=dict[ls[7]];t8=dict[ls[8]]
sum0=sum1=sum2=sum3=sum4=sum5=sum6=sum7=sum8=0
sumr0=sumr1=sumr2=sumr3=sumr4=sumr5=sumr6=sumr7=sumr8=0
for i in range(9,30):
    data=sh2.cell(i,5).value
    for j in range(0,len(t0)):
        if data==t0[j]:
            sum0=sum0+sh2.cell(i,6).value
            sumr0=sumr0+sh2.cell(i,7).value
    for j in range(0,len(t1)):
        if data==t1[j]:
            sum1=sum1+sh2.cell(i,6).value
            sumr1=sumr1+sh2.cell(i,7).value
    for j in range(0,len(t2)):
        if data==t2[j]:
            sum2=sum2+sh2.cell(i,6).value
            sumr2=sumr2+sh2.cell(i,7).value
    for j in range(0,len(t3)):
        if data==t3[j]:
            sum3=sum3+sh2.cell(i,6).value
            sumr3=sumr3+sh2.cell(i,7).value
    for j in range(0,len(t4)):
        if data==t4[j]:
            sum4=sum4+sh2.cell(i,6).value
            sumr4=sumr4+sh2.cell(i,7).value
    for j in range(0,len(t5)):
        if data==t5[j]:
            sum5=sum5+sh2.cell(i,6).value
            sumr5=sumr5+sh2.cell(i,7).value
    for j in range(0,len(t6)):
        if data==t6[j]:
            sum6=sum6+sh2.cell(i,6).value
            sumr6=sumr6+sh2.cell(i,7).value
    for j in range(0,len(t7)):
        if data==t7[j]:
            sum7=sum7+sh2.cell(i,6).value
            sumr7=sumr7+sh2.cell(i,7).value
    for j in range(0,len(t8)):
        if data==t8[j]:
            sum8=sum8+sh2.cell(i,6).value
            sumr8=sumr8+sh2.cell(i,7).value
list1=[sum0/len(t0),sum1/len(t1),sum2/len(t2),sum3/len(t3),sum4/len(t4),sum5/len(t5),sum6/len(t6),sum7/len(t7),sum8/len(t8)]
list2=[sumr0/len(t0),sumr1/len(t1),sumr2/len(t2),sumr3/len(t3),sumr4/len(t4),sumr5/len(t5),sumr6/len(t6),sumr7/len(t7),sumr8/len(t8)]
sh3=wb.create_sheet("result.xlsx")
col_name=["rank","team","statement",'reasons']
sh3.append(col_name)
for i in range(1,len(list1)+1):
    sh3.cell(i+1,3).value=list1[i-1]
    sh3.cell(i+1,2).value=ls[i-1]
    sh3.cell(i+1,4).value=list2[i-1]
wb.save("Python Assignment.xlsx")
df=pd.read_excel('Python Assignment.xlsx',sheet_name='result.xlsx')
df_sort=df.sort_values(by='statement',ascending=False)
writer=pd.ExcelWriter('Python Assignment.xlsx',engine='openpyxl',mode='a')
df_sort.to_excel(writer,sheet_name="output",index=False)
writer.save()
wb=openpyxl.load_workbook("Python Assignment.xlsx")
wb.remove(wb['result.xlsx'])
wb.save('Python Assignment.xlsx')
sh2=wb['(Input) Rigorbuilder RAW']
sh4=wb.create_sheet("temp")
header=["rank",'name','uid','staements','reasons','avg']
sh4.append(header)
for i in range(2,22):
    sh4.cell(i,2).value=sh2.cell(i+7,4).value
    sh4.cell(i,3).value=sh2.cell(i+7,5).value
    sh4.cell(i,4).value=sh2.cell(i+7,6).value
    sh4.cell(i,5).value=sh2.cell(i+7,7).value
    sh4.cell(i,6).value=(sh2.cell(i+7,6).value+sh2.cell(i+7,7).value)/2
wb.save('Python Assignment.xlsx')
df=pd.read_excel('Python Assignment.xlsx',sheet_name='temp')
df_sort=df.sort_values(by='avg',ascending=False)
writer=pd.ExcelWriter('Python Assignment.xlsx',engine='openpyxl',mode='a')
df_sort.to_excel(writer,sheet_name="output2",index=False)
writer.save()
wb=openpyxl.load_workbook("Python Assignment.xlsx")
sheet=wb['output2']
sheet.delete_cols(6)
wb.remove(wb['temp'])
sheet2=wb['output']
for i in range(2,len(ls)+2):
    sheet2.cell(i,1).value=i-1
for j in range(2,22):
    sheet.cell(j,1).value=j-1
wb.save('Python Assignment.xlsx')
print("Done")
print("-------x-------")